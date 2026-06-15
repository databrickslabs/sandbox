"""Shared Excel generation and inspection helpers for Sprint 5 regression."""
import os
import tempfile
from datetime import datetime
from types import SimpleNamespace

import openpyxl

from tests.regression.conftest import make_all_regression_items
from app.routes.export import build_estimate_excel

# Column indices (1-indexed for openpyxl, matching excel_columns.py layout)
COL_IDX = 1
COL_NAME = 2
COL_TYPE = 3
COL_MODE = 4
COL_CONFIG = 5
COL_SKU = 6
COL_DRIVER = 7
COL_WORKER = 8
COL_NUM_WORKERS = 9
COL_DRIVER_TIER = 10
COL_WORKER_TIER = 11
COL_HOURS = 12
COL_TOKEN_TYPE = 13
COL_TOKEN_QTY = 14
COL_DBU_PER_M = 15
COL_DBU_HR = 16
COL_DBUS_MO = 17
COL_DBU_RATE = 18
COL_DISCOUNT = 19
COL_DBU_RATE_DISC = 20
COL_DBU_COST_L = 21
COL_DBU_COST_D = 22
COL_DRIVER_VM_HR = 23
COL_WORKER_VM_HR = 24
COL_DRIVER_VM_COST = 25
COL_WORKER_VM_COST = 26
COL_TOTAL_VM = 27
COL_TOTAL_L = 28
COL_TOTAL_D = 29
COL_NOTES = 30


def make_estimate(**kw):
    d = dict(
        estimate_name='Sprint 5 Regression Test',
        status='draft', version=1,
        created_at=datetime(2026, 4, 3),
        updated_at=datetime(2026, 4, 3),
    )
    d.update(kw)
    return SimpleNamespace(**d)


def generate_xlsx(line_items=None, cloud='aws', region='us-east-1',
                  tier='PREMIUM'):
    """Generate an Excel workbook and return as openpyxl Workbook."""
    if line_items is None:
        line_items = make_all_regression_items()
    estimate = make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp = f.name
    wb = openpyxl.load_workbook(tmp)
    os.unlink(tmp)
    return wb


def _is_valid_sku(val):
    if not val or not isinstance(val, str):
        return False
    if val.startswith('=') or '_' not in val:
        return False
    if val in ('SKU', 'SKU / Product Type'):
        return False
    return True


def find_all_data_rows(ws):
    """Return 1-indexed row numbers for all data rows (by SKU column)."""
    return [r for r in range(1, ws.max_row + 1)
            if _is_valid_sku(ws.cell(row=r, column=COL_SKU).value)]


def find_row_by_name(ws, name):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=COL_NAME).value == name:
            return r
    return None


def find_rows_by_sku(ws, sku):
    return [r for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=COL_SKU).value == sku]


def find_totals_row(ws):
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and 'TOTALS' in str(val):
            return r
    return None


def find_cost_summary_header(ws):
    """Find the row with 'COST SUMMARY' section header."""
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and 'COST SUMMARY' in str(val):
            return r
    return None


def find_dbu_summary_row(ws):
    """Find the 'Total DBUs/Month:' row."""
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and 'Total DBUs' in str(val):
            return r
    return None
