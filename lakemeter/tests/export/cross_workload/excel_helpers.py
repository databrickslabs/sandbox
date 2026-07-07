"""Shared Excel test helpers for Sprint 10 combined estimate tests."""
import os
import tempfile
from datetime import datetime
from types import SimpleNamespace

import openpyxl

from tests.export.cross_workload.conftest import make_all_nine_items
from app.routes.export import build_estimate_excel

# Column indices (1-indexed for openpyxl)
COL_IDX = 1
COL_NAME = 2
COL_TYPE = 3
COL_MODE = 4
COL_CONFIG = 5
COL_SKU = 6
COL_DRIVER = 7
COL_WORKER = 8
COL_NUM_WORKERS = 9
COL_DRIVER_TIER = 10
COL_WORKER_TIER = 11
COL_HOURS = 12
COL_TOKEN_TYPE = 13
COL_TOKEN_QTY = 14
COL_DBU_PER_M = 15
COL_DBU_HR = 16
COL_DBUS_MO = 17
COL_DBU_RATE = 18
COL_DISCOUNT = 19
COL_DBU_RATE_DISC = 20
COL_DBU_COST_L = 21
COL_DBU_COST_D = 22
COL_DRIVER_VM_HR = 23
COL_WORKER_VM_HR = 24
COL_DRIVER_VM_COST = 25
COL_WORKER_VM_COST = 26
COL_TOTAL_VM = 27
COL_TOTAL_L = 28
COL_TOTAL_D = 29
COL_NOTES = 30

# Expected workload types in order
WORKLOAD_TYPES = [
    "JOBS", "ALL_PURPOSE", "DLT", "DBSQL",
    "MODEL_SERVING", "FMAPI_DATABRICKS", "FMAPI_PROPRIETARY",
    "VECTOR_SEARCH", "LAKEBASE",
]


def make_estimate(**kw):
    d = dict(
        estimate_name='Combined E2E Test',
        status='draft',
        version=1,
        created_at=datetime(2026, 3, 31),
        updated_at=datetime(2026, 3, 31),
    )
    d.update(kw)
    return SimpleNamespace(**d)


def generate_xlsx(line_items=None, cloud='aws', region='us-east-1',
                  tier='PREMIUM'):
    """Generate Excel workbook from line items. Returns openpyxl Workbook."""
    if line_items is None:
        line_items = make_all_nine_items()
    estimate = make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def _is_valid_sku(sku_val):
    """Check if a string looks like a real Databricks SKU product type."""
    if not sku_val or not isinstance(sku_val, str):
        return False
    # Real SKUs contain underscores and are ALL_CAPS_WITH_UNDERSCORES
    # Exclude headers, formula refs, dates, region names
    if sku_val.startswith('='):
        return False
    if '_' not in sku_val:
        return False
    if sku_val in ('SKU', 'SKU / Product Type'):
        return False
    return True


def find_all_data_rows(ws):
    """Find all data rows between table headers and totals row."""
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        sku_val = ws.cell(row=row_idx, column=COL_SKU).value
        if _is_valid_sku(sku_val):
            rows.append(row_idx)
    return rows


def find_rows_by_sku(ws, sku):
    """Find all rows matching a specific SKU."""
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=COL_SKU).value == sku:
            rows.append(row_idx)
    return rows


def find_totals_row(ws):
    """Find the TOTALS row by scanning for 'TOTALS:' label."""
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1).value
        if cell and 'TOTALS' in str(cell):
            return row_idx
    return None


def find_row_by_name(ws, name):
    """Find data row by workload name."""
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=COL_NAME).value == name:
            return row_idx
    return None
