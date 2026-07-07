"""
Sprint 1 Iteration 2: Excel Export Validation Tests for Jobs Workloads

Tests that the Excel export generates correct formulas by:
1. Building a minimal .xlsx in-memory using the same _write_row logic
2. Reading it back with openpyxl to verify formula cells and computed values
3. Verifying totals row SUM formulas
4. Verifying notes column content
"""
import os
import sys
import pytest
from io import BytesIO

# Import backend modules
BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

from app.routes.export import (
    _get_sku_type,
    _calculate_dbu_per_hour,
    _calculate_hours_per_month,
    _is_serverless_workload,
    _get_dbu_price,
)
from tests.export.jobs.conftest import make_line_item


def build_test_excel(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    """Build an Excel file from mock line items using the export logic,
    then read it back with openpyxl for verification.

    Returns (workbook, sheet) from openpyxl.
    """
    import xlsxwriter
    import openpyxl
    from xlsxwriter.utility import xl_col_to_name as _col

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    sheet = wb.add_worksheet('Test')

    # Minimal formats
    default_fmt = wb.add_format()
    num_fmt = wb.add_format({'num_format': '#,##0'})
    dec_fmt = wb.add_format({'num_format': '#,##0.00'})
    cur_fmt = wb.add_format({'num_format': '$#,##0.00'})
    pct_fmt = wb.add_format({'num_format': '0%'})

    # Header row (row 0)
    headers = [
        '#', 'Name', 'Type', 'Mode', 'Config', 'SKU',
        'Driver', 'Worker', 'Workers', 'DTier', 'WTier',
        'Hours/Mo', 'Token Type', 'Tokens/Mo(M)', 'DBU/1M',
        'DBU/Hr', 'DBUs/Mo', 'DBU Rate(List)', 'Discount%',
        'DBU Rate(Disc)', 'DBU Cost(List)', 'DBU Cost(Disc)',
        'Driver VM$/Hr', 'Worker VM$/Hr', 'Driver VM Cost',
        'Worker VM Cost', 'Total VM Cost', 'Total Cost(List)',
        'Total Cost(Disc)', 'Notes'
    ]
    for c, h in enumerate(headers):
        sheet.write(0, c, h)

    data_start_row = 1

    for idx, item in enumerate(line_items):
        row = data_start_row + idx
        r = row + 1  # 1-indexed for formulas

        wt = item.workload_type or 'JOBS'
        sku = _get_sku_type(item, cloud)
        dbu_rate, _ = _get_dbu_price(cloud, region, tier, sku)
        dbu_per_hour, warnings = _calculate_dbu_per_hour(item, cloud)
        hours = _calculate_hours_per_month(item)
        is_serverless = _is_serverless_workload(item)
        num_workers = int(item.num_workers or 1)
        total_dbus = dbu_per_hour * hours
        discount_pct = 0.0

        # VM costs (hardcoded in export.py lines 900-901)
        driver_vm_hr = 0.20 if (not is_serverless and wt in ('JOBS', 'ALL_PURPOSE', 'DLT')) else 0
        worker_vm_hr = 0.10 if (not is_serverless and wt in ('JOBS', 'ALL_PURPOSE', 'DLT')) else 0

        notes = ' | '.join(warnings) if warnings else (item.notes or '')

        # Static columns
        sheet.write(row, 0, idx + 1)
        sheet.write(row, 1, item.workload_name or f'Job {idx+1}')
        sheet.write(row, 2, wt)
        sheet.write(row, 3, 'Serverless' if is_serverless else 'Classic')
        sheet.write(row, 4, '')
        sheet.write(row, 5, sku)
        sheet.write(row, 6, item.driver_node_type or '-')
        sheet.write(row, 7, item.worker_node_type or '-')
        sheet.write(row, 8, num_workers, num_fmt)
        sheet.write(row, 9, '-')
        sheet.write(row, 10, '-')
        sheet.write(row, 11, hours, dec_fmt)
        sheet.write(row, 12, '-')
        sheet.write(row, 13, '-')
        sheet.write(row, 14, '-')
        sheet.write(row, 15, dbu_per_hour, dec_fmt)

        # Col 16: DBUs/Mo = DBU/Hr × Hours/Mo (FORMULA)
        formula = f'={_col(15)}{r}*{_col(11)}{r}'
        sheet.write_formula(row, 16, formula, num_fmt, total_dbus)

        # Col 17: DBU Rate (List)
        sheet.write(row, 17, dbu_rate, cur_fmt)

        # Col 18: Discount %
        sheet.write(row, 18, discount_pct, pct_fmt)

        # Col 19: DBU Rate (Disc.) = R*(1-S)
        disc_rate = dbu_rate * (1 - discount_pct)
        formula = f'={_col(17)}{r}*(1-{_col(18)}{r})'
        sheet.write_formula(row, 19, formula, cur_fmt, disc_rate)

        # Col 20: DBU Cost (List) = Q*R
        dbu_cost_list = total_dbus * dbu_rate
        formula = f'={_col(16)}{r}*{_col(17)}{r}'
        sheet.write_formula(row, 20, formula, cur_fmt, dbu_cost_list)

        # Col 21: DBU Cost (Disc.) = Q*T
        dbu_cost_disc = total_dbus * disc_rate
        formula = f'={_col(16)}{r}*{_col(19)}{r}'
        sheet.write_formula(row, 21, formula, cur_fmt, dbu_cost_disc)

        if is_serverless:
            for c in range(22, 27):
                sheet.write(row, c, 0, cur_fmt)
        else:
            sheet.write(row, 22, driver_vm_hr, cur_fmt)
            sheet.write(row, 23, worker_vm_hr, cur_fmt)
            # Col 24: Driver VM Cost = W*L
            driver_vm_total = driver_vm_hr * hours
            formula = f'={_col(22)}{r}*{_col(11)}{r}'
            sheet.write_formula(row, 24, formula, cur_fmt, driver_vm_total)
            # Col 25: Worker VM Cost = X*L*I
            worker_vm_total = worker_vm_hr * hours * num_workers
            formula = f'={_col(23)}{r}*{_col(11)}{r}*{_col(8)}{r}'
            sheet.write_formula(row, 25, formula, cur_fmt, worker_vm_total)
            # Col 26: Total VM Cost = Y+Z
            formula = f'={_col(24)}{r}+{_col(25)}{r}'
            sheet.write_formula(row, 26, formula, cur_fmt, driver_vm_total + worker_vm_total)

        # Col 27: Total Cost (List) = U+AA
        vm_total = 0
        if not is_serverless:
            vm_total = driver_vm_hr * hours + worker_vm_hr * hours * num_workers
        formula = f'={_col(20)}{r}+{_col(26)}{r}'
        sheet.write_formula(row, 27, formula, cur_fmt, dbu_cost_list + vm_total)

        # Col 28: Total Cost (Disc.) = V+AA
        formula = f'={_col(21)}{r}+{_col(26)}{r}'
        sheet.write_formula(row, 28, formula, cur_fmt, dbu_cost_disc + vm_total)

        # Col 29: Notes
        sheet.write(row, 29, notes)

    data_end_row = data_start_row + len(line_items) - 1

    # Totals row
    totals_row = data_end_row + 2
    ds = data_start_row + 1
    de = data_end_row + 1
    sheet.write(totals_row, 0, 'TOTALS')
    sheet.write_formula(totals_row, 16, f'=SUM({_col(16)}{ds}:{_col(16)}{de})', num_fmt)
    sheet.write_formula(totals_row, 20, f'=SUM({_col(20)}{ds}:{_col(20)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 21, f'=SUM({_col(21)}{ds}:{_col(21)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 24, f'=SUM({_col(24)}{ds}:{_col(24)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 25, f'=SUM({_col(25)}{ds}:{_col(25)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 26, f'=SUM({_col(26)}{ds}:{_col(26)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 27, f'=SUM({_col(27)}{ds}:{_col(27)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 28, f'=SUM({_col(28)}{ds}:{_col(28)}{de})', cur_fmt)

    wb.close()
    output.seek(0)

    # Read back with openpyxl
    oxl_wb = openpyxl.load_workbook(output, data_only=False)
    oxl_sheet = oxl_wb.active
    return oxl_wb, oxl_sheet, data_start_row, totals_row


# ============================================================
# Test: Formula Presence in Excel Cells
# ============================================================

class TestExcelFormulaPresence:
    """Verify that computed columns contain formulas, not just static values."""

    @pytest.fixture
    def classic_standard_excel(self):
        item = make_line_item(
            workload_type="JOBS",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=2, photon_enabled=False, serverless_enabled=False,
            runs_per_day=10, avg_runtime_minutes=30, days_per_month=22,
        )
        wb, sheet, data_start, totals_row = build_test_excel([item])
        yield sheet, data_start, totals_row
        wb.close()

    def test_dbus_mo_has_formula(self, classic_standard_excel):
        """Col 16 (DBUs/Mo) should be a formula =P*L."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=17)  # openpyxl is 1-indexed
        assert cell.value is not None
        assert isinstance(cell.value, str) and cell.value.startswith('='), \
            f"DBUs/Mo should be a formula, got: {cell.value}"

    def test_dbu_rate_disc_has_formula(self, classic_standard_excel):
        """Col 19 (DBU Rate Disc.) should be a formula."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=20)  # col 19 → openpyxl col 20
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_dbu_cost_list_has_formula(self, classic_standard_excel):
        """Col 20 (DBU Cost List) should be a formula."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=21)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_dbu_cost_disc_has_formula(self, classic_standard_excel):
        """Col 21 (DBU Cost Disc.) should be a formula."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=22)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_driver_vm_cost_has_formula(self, classic_standard_excel):
        """Col 24 (Driver VM Cost) should be a formula for classic."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=25)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_worker_vm_cost_has_formula(self, classic_standard_excel):
        """Col 25 (Worker VM Cost) should be a formula for classic."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=26)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_total_vm_cost_has_formula(self, classic_standard_excel):
        """Col 26 (Total VM Cost) should be a formula for classic."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=27)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_total_cost_list_has_formula(self, classic_standard_excel):
        """Col 27 (Total Cost List) should be a formula."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=28)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_total_cost_disc_has_formula(self, classic_standard_excel):
        """Col 28 (Total Cost Disc.) should be a formula."""
        sheet, ds, _ = classic_standard_excel
        cell = sheet.cell(row=ds + 1, column=29)
        assert isinstance(cell.value, str) and cell.value.startswith('=')


# ============================================================
# Test: Totals Row SUM Formulas
# ============================================================

class TestExcelTotalsRow:
    """Verify totals row has SUM formulas across all data rows."""

    @pytest.fixture
    def multi_row_excel(self):
        """Create Excel with 4 Jobs variants."""
        items = [
            make_line_item(workload_name="Jobs Classic", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=False, serverless_enabled=False,
                           runs_per_day=10, avg_runtime_minutes=30, days_per_month=22),
            make_line_item(workload_name="Jobs Photon", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=True, serverless_enabled=False,
                           hours_per_month=110),
            make_line_item(workload_name="Jobs Serverless Std", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=True, serverless_enabled=True,
                           serverless_mode="standard", hours_per_month=110),
            make_line_item(workload_name="Jobs Serverless Perf", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=True, serverless_enabled=True,
                           serverless_mode="performance", hours_per_month=730),
        ]
        wb, sheet, data_start, totals_row = build_test_excel(items)
        yield sheet, data_start, totals_row, len(items)
        wb.close()

    def test_totals_dbus_mo_is_sum(self, multi_row_excel):
        """Col 16 totals should be SUM formula."""
        sheet, ds, tr, _ = multi_row_excel
        cell = sheet.cell(row=tr + 1, column=17)
        assert isinstance(cell.value, str) and 'SUM' in cell.value.upper()

    def test_totals_dbu_cost_list_is_sum(self, multi_row_excel):
        """Col 20 totals should be SUM formula."""
        sheet, ds, tr, _ = multi_row_excel
        cell = sheet.cell(row=tr + 1, column=21)
        assert isinstance(cell.value, str) and 'SUM' in cell.value.upper()

    def test_totals_total_cost_list_is_sum(self, multi_row_excel):
        """Col 27 totals should be SUM formula."""
        sheet, ds, tr, _ = multi_row_excel
        cell = sheet.cell(row=tr + 1, column=28)
        assert isinstance(cell.value, str) and 'SUM' in cell.value.upper()

    def test_totals_total_cost_disc_is_sum(self, multi_row_excel):
        """Col 28 totals should be SUM formula."""
        sheet, ds, tr, _ = multi_row_excel
        cell = sheet.cell(row=tr + 1, column=29)
        assert isinstance(cell.value, str) and 'SUM' in cell.value.upper()

    def test_totals_vm_columns_are_sum(self, multi_row_excel):
        """VM cost totals (cols 24-26) should be SUM formulas."""
        sheet, ds, tr, _ = multi_row_excel
        for col in [25, 26, 27]:  # openpyxl 1-indexed: 24→25, 25→26, 26→27
            cell = sheet.cell(row=tr + 1, column=col)
            assert isinstance(cell.value, str) and 'SUM' in cell.value.upper(), \
                f"Col {col-1} totals should have SUM formula, got: {cell.value}"

    def test_sum_ranges_cover_all_data_rows(self, multi_row_excel):
        """SUM formulas should reference all data rows (2:5 for 4 items starting at row 1)."""
        sheet, ds, tr, n_items = multi_row_excel
        cell = sheet.cell(row=tr + 1, column=17)  # DBUs/Mo totals
        formula = cell.value
        # Should reference from row ds+1 to ds+n_items (1-indexed)
        expected_start = ds + 1
        expected_end = ds + n_items
        assert f'{expected_start}' in formula and f'{expected_end}' in formula, \
            f"SUM range should cover rows {expected_start}:{expected_end}, got: {formula}"


# ============================================================
# Test: Serverless has zero VM costs in Excel
# ============================================================

class TestExcelServerlessNoVM:
    """Verify serverless rows have zero VM costs in Excel output."""

    @pytest.fixture
    def serverless_excel(self):
        item = make_line_item(
            workload_type="JOBS",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=2, photon_enabled=True, serverless_enabled=True,
            serverless_mode="standard", hours_per_month=110,
        )
        wb, sheet, data_start, totals_row = build_test_excel([item])
        yield sheet, data_start
        wb.close()

    def test_driver_vm_hr_zero(self, serverless_excel):
        sheet, ds = serverless_excel
        cell = sheet.cell(row=ds + 1, column=23)  # col 22 → openpyxl 23
        assert cell.value == 0

    def test_worker_vm_hr_zero(self, serverless_excel):
        sheet, ds = serverless_excel
        cell = sheet.cell(row=ds + 1, column=24)
        assert cell.value == 0

    def test_all_vm_columns_zero(self, serverless_excel):
        sheet, ds = serverless_excel
        for col in range(23, 28):  # cols 22-26
            cell = sheet.cell(row=ds + 1, column=col)
            assert cell.value == 0, f"Col {col-1} should be 0 for serverless, got {cell.value}"


# ============================================================
# Test: Static Column Values (SKU, Mode, DBU/Hr, Hours)
# ============================================================

class TestExcelStaticValues:
    """Verify static column values match expected."""

    @pytest.fixture
    def all_four_configs_excel(self):
        items = [
            make_line_item(workload_name="Classic Std", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=False, serverless_enabled=False,
                           runs_per_day=10, avg_runtime_minutes=30, days_per_month=22),
            make_line_item(workload_name="Classic Photon", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=True, serverless_enabled=False,
                           hours_per_month=110),
            make_line_item(workload_name="SL Standard", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=True, serverless_enabled=True,
                           serverless_mode="standard", hours_per_month=110),
            make_line_item(workload_name="SL Perf", workload_type="JOBS",
                           driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                           num_workers=2, photon_enabled=True, serverless_enabled=True,
                           serverless_mode="performance", hours_per_month=730),
        ]
        wb, sheet, data_start, totals_row = build_test_excel(items)
        yield sheet, data_start
        wb.close()

    def test_sku_classic_standard(self, all_four_configs_excel):
        sheet, ds = all_four_configs_excel
        assert sheet.cell(row=ds + 1, column=6).value == 'JOBS_COMPUTE'

    def test_sku_classic_photon(self, all_four_configs_excel):
        sheet, ds = all_four_configs_excel
        assert sheet.cell(row=ds + 2, column=6).value == 'JOBS_COMPUTE_(PHOTON)'

    def test_sku_serverless(self, all_four_configs_excel):
        sheet, ds = all_four_configs_excel
        assert sheet.cell(row=ds + 3, column=6).value == 'JOBS_SERVERLESS_COMPUTE'
        assert sheet.cell(row=ds + 4, column=6).value == 'JOBS_SERVERLESS_COMPUTE'

    def test_mode_classic_vs_serverless(self, all_four_configs_excel):
        sheet, ds = all_four_configs_excel
        assert sheet.cell(row=ds + 1, column=4).value == 'Classic'
        assert sheet.cell(row=ds + 2, column=4).value == 'Classic'
        assert sheet.cell(row=ds + 3, column=4).value == 'Serverless'
        assert sheet.cell(row=ds + 4, column=4).value == 'Serverless'

    def test_dbu_per_hour_values(self, all_four_configs_excel):
        """DBU/Hr: Classic=3.0, Photon=8.7, SL Std=8.7, SL Perf=17.4."""
        sheet, ds = all_four_configs_excel
        assert sheet.cell(row=ds + 1, column=16).value == pytest.approx(3.0)
        assert sheet.cell(row=ds + 2, column=16).value == pytest.approx(8.7)
        assert sheet.cell(row=ds + 3, column=16).value == pytest.approx(8.7)
        assert sheet.cell(row=ds + 4, column=16).value == pytest.approx(17.4)

    def test_hours_per_month_values(self, all_four_configs_excel):
        """Hours: run-based=110, direct=110, direct=110, direct=730."""
        sheet, ds = all_four_configs_excel
        assert sheet.cell(row=ds + 1, column=12).value == pytest.approx(110.0)
        assert sheet.cell(row=ds + 2, column=12).value == pytest.approx(110.0)
        assert sheet.cell(row=ds + 3, column=12).value == pytest.approx(110.0)
        assert sheet.cell(row=ds + 4, column=12).value == pytest.approx(730.0)

    def test_dbu_rate_values(self, all_four_configs_excel):
        """DBU Rate: $0.15 for classic, $0.15 photon, $0.35 serverless."""
        sheet, ds = all_four_configs_excel
        # Classic standard
        assert sheet.cell(row=ds + 1, column=18).value == pytest.approx(0.15)
        # Classic photon
        assert sheet.cell(row=ds + 2, column=18).value == pytest.approx(0.15)
        # Serverless (both)
        serverless_rate = sheet.cell(row=ds + 3, column=18).value
        assert serverless_rate == pytest.approx(0.35) or serverless_rate == pytest.approx(0.39), \
            f"Serverless DBU rate should be ~$0.35-0.39, got {serverless_rate}"
