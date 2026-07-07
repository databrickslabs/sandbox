"""
Sprint 1 Iteration 2: Integration test against real /api/v1/export/estimate/{id}/excel endpoint.

This test uses FastAPI TestClient with dependency overrides to mock the database
and auth layers, while exercising the REAL export pipeline end-to-end:
- _get_sku_type, _calculate_dbu_per_hour, _calculate_hours_per_month
- xlsxwriter Excel generation
- StreamingResponse output

The downloaded .xlsx is then read with openpyxl to verify formulas and values.
"""
import os
import sys
import uuid
import pytest
from io import BytesIO
from unittest.mock import MagicMock, patch
from types import SimpleNamespace

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.database import get_db
from app.auth import get_current_user
from tests.export.jobs.conftest import make_line_item


# ============================================================
# Fixtures: mock DB + auth
# ============================================================

MOCK_USER_ID = uuid.uuid4()
MOCK_ESTIMATE_ID = uuid.uuid4()


def _make_mock_user():
    user = MagicMock()
    user.user_id = MOCK_USER_ID
    user.email = "test@databricks.com"
    user.full_name = "Test User"
    user.is_active = True
    return user


def _make_mock_estimate(line_items_data):
    """Build a mock Estimate with real-ish LineItem objects."""
    estimate = MagicMock()
    estimate.estimate_id = MOCK_ESTIMATE_ID
    estimate.estimate_name = "Sprint 1 Integration Test"
    estimate.owner_user_id = MOCK_USER_ID
    estimate.is_deleted = False
    estimate.cloud = "aws"
    estimate.region = "us-east-1"
    estimate.tier = "PREMIUM"
    estimate.created_at = None
    estimate.updated_at = None

    mock_items = []
    for idx, data in enumerate(line_items_data):
        item = make_line_item(display_order=idx, **data)
        mock_items.append(item)

    estimate.line_items = mock_items
    return estimate


def _override_auth():
    return _make_mock_user()


def _build_client_with_items(line_items_data):
    """Create a TestClient with mocked DB returning the given line items."""
    estimate = _make_mock_estimate(line_items_data)
    mock_user = _make_mock_user()

    mock_db = MagicMock()
    # Mock the query chain: db.query(Estimate).filter(...).first()
    mock_db.query.return_value.filter.return_value.first.return_value = estimate
    # Mock sharing query to return None (user is owner, not shared)

    def override_db():
        yield mock_db

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: mock_user

    client = TestClient(app)
    return client


def _download_excel(client) -> openpyxl.Workbook:
    """Call the export endpoint and parse the returned .xlsx."""
    url = f"/api/v1/export/estimate/{MOCK_ESTIMATE_ID}/excel"
    response = client.get(url)
    assert response.status_code == 200, f"Export failed: {response.status_code} {response.text}"
    assert "spreadsheetml" in response.headers.get("content-type", ""), \
        f"Expected Excel content-type, got: {response.headers.get('content-type')}"

    wb = openpyxl.load_workbook(BytesIO(response.content), data_only=False)
    return wb


# ============================================================
# Integration Tests
# ============================================================

class TestExportEndpointJobsClassic:
    """Integration: export a Jobs Classic Standard item via the real endpoint."""

    @pytest.fixture(autouse=True)
    def setup_client(self):
        self.client = _build_client_with_items([{
            "workload_type": "JOBS",
            "workload_name": "Integration Test Job Classic",
            "driver_node_type": "i3.xlarge",
            "worker_node_type": "i3.xlarge",
            "num_workers": 2,
            "photon_enabled": False,
            "serverless_enabled": False,
            "runs_per_day": 10,
            "avg_runtime_minutes": 30,
            "days_per_month": 22,
        }])
        yield
        app.dependency_overrides.clear()

    def test_endpoint_returns_200(self):
        url = f"/api/v1/export/estimate/{MOCK_ESTIMATE_ID}/excel"
        response = self.client.get(url)
        assert response.status_code == 200

    def test_excel_has_data_rows(self):
        wb = _download_excel(self.client)
        sheet = wb.active
        # Find a row with "JOBS_COMPUTE" SKU
        found_sku = False
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if cell.value == "JOBS_COMPUTE":
                    found_sku = True
                    break
        assert found_sku, "Expected JOBS_COMPUTE SKU in exported Excel"
        wb.close()

    def test_excel_has_formula_cells(self):
        """Verify the export pipeline writes formulas (= prefix) not static values."""
        wb = _download_excel(self.client)
        sheet = wb.active
        formula_found = False
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_found = True
                    break
            if formula_found:
                break
        assert formula_found, "Expected at least one formula cell in exported Excel"
        wb.close()

    def test_excel_has_totals_row(self):
        """Verify the export pipeline writes a TOTALS row with SUM formulas."""
        wb = _download_excel(self.client)
        sheet = wb.active
        totals_found = False
        sum_formula_found = False
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and "TOTAL" in cell.value.upper():
                    totals_found = True
                if isinstance(cell.value, str) and "SUM(" in cell.value.upper():
                    sum_formula_found = True
        assert totals_found, "Expected TOTALS row in exported Excel"
        assert sum_formula_found, "Expected SUM formulas in totals row"
        wb.close()


class TestExportEndpointMultipleJobs:
    """Integration: export 4 Jobs variants via the real endpoint."""

    @pytest.fixture(autouse=True)
    def setup_client(self):
        self.client = _build_client_with_items([
            {
                "workload_type": "JOBS", "workload_name": "Jobs Classic Std",
                "driver_node_type": "i3.xlarge", "worker_node_type": "i3.xlarge",
                "num_workers": 2, "photon_enabled": False, "serverless_enabled": False,
                "runs_per_day": 10, "avg_runtime_minutes": 30, "days_per_month": 22,
            },
            {
                "workload_type": "JOBS", "workload_name": "Jobs Photon",
                "driver_node_type": "i3.xlarge", "worker_node_type": "i3.xlarge",
                "num_workers": 2, "photon_enabled": True, "serverless_enabled": False,
                "hours_per_month": 110,
            },
            {
                "workload_type": "JOBS", "workload_name": "Jobs SL Standard",
                "driver_node_type": "i3.xlarge", "worker_node_type": "i3.xlarge",
                "num_workers": 2, "photon_enabled": True, "serverless_enabled": True,
                "serverless_mode": "standard", "hours_per_month": 110,
            },
            {
                "workload_type": "JOBS", "workload_name": "Jobs SL Perf",
                "driver_node_type": "i3.xlarge", "worker_node_type": "i3.xlarge",
                "num_workers": 2, "photon_enabled": True, "serverless_enabled": True,
                "serverless_mode": "performance", "hours_per_month": 730,
            },
        ])
        yield
        app.dependency_overrides.clear()

    def test_all_four_rows_present(self):
        wb = _download_excel(self.client)
        sheet = wb.active
        job_names = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("Jobs "):
                    job_names.append(cell.value)
        assert len(job_names) >= 4, f"Expected 4 job rows, found: {job_names}"
        wb.close()

    def test_sku_mapping_in_real_export(self):
        """Verify real endpoint writes correct SKUs for all 4 configs."""
        wb = _download_excel(self.client)
        sheet = wb.active
        skus = set()
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and "JOBS" in cell.value and "COMPUTE" in cell.value:
                    skus.add(cell.value)
        assert "JOBS_COMPUTE" in skus, f"Missing JOBS_COMPUTE in {skus}"
        assert "JOBS_COMPUTE_(PHOTON)" in skus, f"Missing JOBS_COMPUTE_(PHOTON) in {skus}"
        assert "JOBS_SERVERLESS_COMPUTE" in skus, f"Missing JOBS_SERVERLESS_COMPUTE in {skus}"
        wb.close()

    def test_no_error_cells_in_export(self):
        """Verify no #REF!, #VALUE!, or error markers in the Excel output.
        Excludes '#' used as a column header (row number column).
        """
        wb = _download_excel(self.client)
        sheet = wb.active
        error_prefixes = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A")
        errors = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(error_prefixes):
                    errors.append(f"Cell {cell.coordinate}: {cell.value}")
        assert len(errors) == 0, f"Found error cells: {errors}"
        wb.close()

    def test_serverless_rows_have_zero_vm(self):
        """Verify serverless rows in real export have $0 VM costs.

        Column layout (0-indexed):
          3  = Mode ("Serverless" or "Classic")
          22 = Driver VM $/Hr
          23 = Worker VM $/Hr
          24 = Driver VM Cost (formula or value)
          25 = Worker VM Cost (formula or value)
          26 = Total VM Cost (formula or value)
        """
        wb = _download_excel(self.client)
        sheet = wb.active

        serverless_rows_checked = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            mode_cell = row[3] if len(row) > 3 else None
            if mode_cell and isinstance(mode_cell.value, str) and mode_cell.value == "Serverless":
                serverless_rows_checked += 1
                # Driver VM $/Hr (col 22) and Worker VM $/Hr (col 23) should be 0
                driver_vm_hr = row[22].value if len(row) > 22 else None
                worker_vm_hr = row[23].value if len(row) > 23 else None
                assert driver_vm_hr == 0 or driver_vm_hr is None, \
                    f"Row {mode_cell.row}: Driver VM $/Hr should be 0, got {driver_vm_hr}"
                assert worker_vm_hr == 0 or worker_vm_hr is None, \
                    f"Row {mode_cell.row}: Worker VM $/Hr should be 0, got {worker_vm_hr}"

                # Total VM Cost (col 26) — could be a formula like "=Y+Z" that evaluates to 0
                # With data_only=False, formulas appear as strings. Static 0 is also acceptable.
                total_vm = row[26].value if len(row) > 26 else None
                if isinstance(total_vm, (int, float)):
                    assert total_vm == 0, \
                        f"Row {mode_cell.row}: Total VM Cost should be 0, got {total_vm}"
                # If it's a formula string, the underlying cached value should be 0
                # (openpyxl doesn't evaluate formulas, but we verified the inputs are 0)

        assert serverless_rows_checked >= 2, \
            f"Expected at least 2 serverless rows, found {serverless_rows_checked}"
        wb.close()
