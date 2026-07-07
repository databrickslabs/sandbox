"""Shared Excel test helpers for Vector Search export tests."""
import os
import tempfile
from datetime import datetime
from types import SimpleNamespace

import openpyxl

from tests.export.vector_search.conftest import make_line_item
from app.routes.export import build_estimate_excel

# Column indices (1-indexed for openpyxl)
COL_TYPE = 3
COL_MODE = 4
COL_CONFIG = 5
COL_SKU = 6
COL_HOURS = 12
COL_TOKEN_TYPE = 13
COL_DBU_HR = 16
COL_DBUS_MO = 17
COL_DBU_RATE = 18
COL_DISCOUNT = 19
COL_DBU_RATE_DISC = 20
COL_DBU_COST_L = 21
COL_DBU_COST_D = 22
COL_TOTAL_L = 28
COL_TOTAL_D = 29
COL_NOTES = 30


def make_estimate(**kw):
    d = dict(estimate_name='Vector Search E2E', status='draft', version=1,
             created_at=datetime(2026, 3, 31), updated_at=datetime(2026, 3, 31))
    d.update(kw)
    return SimpleNamespace(**d)


def generate_xlsx(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    estimate = make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def find_data_rows(ws, sku_filter=None):
    """Find all data rows. If sku_filter given, only rows with that SKU."""
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        sku_val = ws.cell(row=row_idx, column=6).value
        if sku_filter:
            if sku_val == sku_filter:
                rows.append(row_idx)
        else:
            if sku_val and isinstance(sku_val, str) and sku_val not in (
                    '', '-', 'SKU', 'SKU / Product Type'):
                rows.append(row_idx)
    return rows


def find_vs_compute_row(ws):
    """Find Vector Search compute row (SERVERLESS_REAL_TIME_INFERENCE)."""
    rows = find_data_rows(ws, 'SERVERLESS_REAL_TIME_INFERENCE')
    return rows[0] if rows else None


def find_storage_row(ws):
    """Find DATABRICKS_STORAGE row."""
    rows = find_data_rows(ws, 'DATABRICKS_STORAGE')
    return rows[0] if rows else None
