"""
Sprint 4: Integration test for multiple DBSQL types (Classic + Pro + Serverless).

Uses FastAPI TestClient with mocked DB/auth to verify all three warehouse types
can be exported together with correct SKUs, formulas, and totals.
"""
import os
import sys
import uuid
import pytest
from io import BytesIO
from unittest.mock import MagicMock

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.database import get_db
from app.auth import get_current_user
from tests.export.dbsql.conftest import make_line_item


MOCK_USER_ID = uuid.uuid4()
MOCK_ESTIMATE_ID = uuid.uuid4()


def _make_mock_user():
    user = MagicMock()
    user.user_id = MOCK_USER_ID
    user.email = "test@databricks.com"
    user.full_name = "Test User"
    user.is_active = True
    return user


def _make_mock_estimate(line_items_data):
    estimate = MagicMock()
    estimate.estimate_id = MOCK_ESTIMATE_ID
    estimate.estimate_name = "Sprint 4 Multi-Type Test"
    estimate.owner_user_id = MOCK_USER_ID
    estimate.is_deleted = False
    estimate.cloud = "aws"
    estimate.region = "us-east-1"
    estimate.tier = "PREMIUM"
    estimate.created_at = None
    estimate.updated_at = None
    estimate.line_items = [
        make_line_item(display_order=i, **d) for i, d in enumerate(line_items_data)
    ]
    return estimate


def _build_client_with_items(line_items_data):
    estimate = _make_mock_estimate(line_items_data)
    mock_db = MagicMock()
    mock_db.query.return_value.filter.return_value.first.return_value = estimate

    def override_db():
        yield mock_db

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: _make_mock_user()
    return TestClient(app)


def _download_excel(client) -> openpyxl.Workbook:
    url = f"/api/v1/export/estimate/{MOCK_ESTIMATE_ID}/excel"
    resp = client.get(url)
    assert resp.status_code == 200
    return openpyxl.load_workbook(BytesIO(resp.content), data_only=False)


class TestDBSQLMultiTypeIntegration:
    """Integration: export Classic, Pro, Serverless DBSQL via real endpoint."""

    @pytest.fixture(autouse=True)
    def setup_client(self):
        self.client = _build_client_with_items([
            {
                "workload_type": "DBSQL", "workload_name": "Classic Small",
                "dbsql_warehouse_type": "CLASSIC", "dbsql_warehouse_size": "Small",
                "dbsql_num_clusters": 1, "hours_per_month": 100,
            },
            {
                "workload_type": "DBSQL", "workload_name": "Pro Medium 2x",
                "dbsql_warehouse_type": "PRO", "dbsql_warehouse_size": "Medium",
                "dbsql_num_clusters": 2, "hours_per_month": 200,
            },
            {
                "workload_type": "DBSQL", "workload_name": "SL Large",
                "dbsql_warehouse_type": "SERVERLESS", "dbsql_warehouse_size": "Large",
                "dbsql_num_clusters": 1, "hours_per_month": 730,
            },
            {
                "workload_type": "DBSQL", "workload_name": "SL 4X-Large",
                "dbsql_warehouse_type": "SERVERLESS", "dbsql_warehouse_size": "4X-Large",
                "dbsql_num_clusters": 1, "hours_per_month": 730,
            },
        ])
        yield
        app.dependency_overrides.clear()

    def test_all_four_rows_present(self):
        wb = _download_excel(self.client)
        names = []
        for row in wb.active.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value in (
                    "Classic Small", "Pro Medium 2x", "SL Large", "SL 4X-Large"
                ):
                    names.append(cell.value)
        assert len(names) == 4, f"Expected 4 rows, found: {names}"
        wb.close()

    def test_all_three_sku_types_present(self):
        wb = _download_excel(self.client)
        skus = set()
        for row in wb.active.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and 'SQL' in cell.value and 'COMPUTE' in cell.value:
                    skus.add(cell.value)
        assert 'SQL_COMPUTE' in skus
        assert 'SQL_PRO_COMPUTE' in skus
        assert 'SERVERLESS_SQL_COMPUTE' in skus
        wb.close()

    def test_has_formula_cells(self):
        wb = _download_excel(self.client)
        found = any(
            isinstance(c.value, str) and c.value.startswith("=")
            for row in wb.active.iter_rows(values_only=False)
            for c in row
        )
        assert found
        wb.close()

    def test_has_totals_with_sum(self):
        wb = _download_excel(self.client)
        found = any(
            isinstance(c.value, str) and "SUM(" in c.value.upper()
            for row in wb.active.iter_rows(values_only=False)
            for c in row
        )
        assert found
        wb.close()

    def test_no_error_cells(self):
        wb = _download_excel(self.client)
        error_prefixes = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A")
        errors = [
            f"{c.coordinate}: {c.value}"
            for row in wb.active.iter_rows(values_only=False)
            for c in row
            if isinstance(c.value, str) and c.value.startswith(error_prefixes)
        ]
        assert len(errors) == 0, f"Errors: {errors}"
        wb.close()
