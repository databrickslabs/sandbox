"""
Sprint 4: DBSQL Excel export formula verification tests.

Builds an Excel workbook using the real export pipeline and reads it back
with openpyxl to verify formulas, row structure, and computed values.
"""
import os
import sys
import pytest
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from app.routes.export.excel_builder import build_estimate_excel
from tests.export.dbsql.conftest import make_line_item


def _build_excel(line_items_data, cloud='aws', region='us-east-1', tier='PREMIUM'):
    """Build an Excel workbook from line item configs and return openpyxl wb."""
    estimate = MagicMock()
    estimate.estimate_name = "Sprint 4 DBSQL Excel Test"
    estimate.status = "draft"
    estimate.version = 1
    estimate.created_at = None
    estimate.updated_at = None

    items = [make_line_item(display_order=i, **data) for i, data in enumerate(line_items_data)]
    output = build_estimate_excel(estimate, items, cloud, region, tier)
    wb = openpyxl.load_workbook(BytesIO(output.read()), data_only=False)
    return wb


def _find_data_rows(sheet, sku_prefix='SQL'):
    """Find rows containing DBSQL SKU values."""
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and sku_prefix in cell.value and 'COMPUTE' in cell.value:
                rows.append(row)
                break
    return rows


class TestDBSQLExcelFormulas:
    """Verify DBSQL rows have formula cells, not static values."""

    @pytest.fixture(autouse=True)
    def setup_wb(self):
        self.wb = _build_excel([{
            "workload_type": "DBSQL",
            "workload_name": "DBSQL Serverless Small",
            "dbsql_warehouse_type": "SERVERLESS",
            "dbsql_warehouse_size": "Small",
            "dbsql_num_clusters": 1,
            "hours_per_month": 100,
        }])
        yield
        self.wb.close()

    def test_dbus_month_is_formula(self):
        """Col 16 (DBUs/Mo) should be a formula =P*L."""
        sheet = self.wb.active
        rows = _find_data_rows(sheet, 'SERVERLESS_SQL')
        assert len(rows) >= 1, "No SERVERLESS_SQL_COMPUTE row found"
        cell = rows[0][16]
        assert isinstance(cell.value, str) and cell.value.startswith('='), \
            f"Col 16 should be formula, got: {cell.value}"

    def test_dbu_rate_disc_is_formula(self):
        """Col 19 (DBU Rate Discounted) should be a formula."""
        sheet = self.wb.active
        rows = _find_data_rows(sheet, 'SERVERLESS_SQL')
        cell = rows[0][19]
        assert isinstance(cell.value, str) and cell.value.startswith('='), \
            f"Col 19 should be formula, got: {cell.value}"

    def test_dbu_cost_list_is_formula(self):
        """Col 20 (DBU Cost List) should be a formula."""
        sheet = self.wb.active
        rows = _find_data_rows(sheet, 'SERVERLESS_SQL')
        cell = rows[0][20]
        assert isinstance(cell.value, str) and cell.value.startswith('='), \
            f"Col 20 should be formula, got: {cell.value}"

    def test_total_cost_is_formula(self):
        """Col 27 (Total Cost List) should be a formula."""
        sheet = self.wb.active
        rows = _find_data_rows(sheet, 'SERVERLESS_SQL')
        cell = rows[0][27]
        assert isinstance(cell.value, str) and cell.value.startswith('='), \
            f"Col 27 should be formula, got: {cell.value}"


class TestDBSQLExcelServerlessModeColumn:
    """Serverless DBSQL rows must show 'Serverless' in col 3."""

    @pytest.fixture(autouse=True)
    def setup_wb(self):
        self.wb = _build_excel([{
            "workload_type": "DBSQL",
            "dbsql_warehouse_type": "SERVERLESS",
            "dbsql_warehouse_size": "Small",
            "hours_per_month": 100,
        }])
        yield
        self.wb.close()

    def test_mode_column_shows_serverless(self):
        sheet = self.wb.active
        rows = _find_data_rows(sheet, 'SERVERLESS_SQL')
        assert len(rows) >= 1
        mode_val = rows[0][3].value
        assert mode_val == "Serverless", f"Expected 'Serverless', got '{mode_val}'"

    def test_vm_cols_are_dash_or_zero(self):
        """Serverless rows have '-' for VM config cols and 0 for VM cost cols."""
        sheet = self.wb.active
        rows = _find_data_rows(sheet, 'SERVERLESS_SQL')
        row = rows[0]
        for c in range(6, 11):
            assert row[c].value == '-', f"Col {c} should be '-', got {row[c].value}"
        for c in range(22, 27):
            val = row[c].value
            assert val == 0 or val is None, f"Col {c} should be 0, got {val}"


class TestDBSQLExcelMultipleTypes:
    """Export with Classic, Pro, and Serverless DBSQL rows."""

    @pytest.fixture(autouse=True)
    def setup_wb(self):
        self.wb = _build_excel([
            {
                "workload_type": "DBSQL", "workload_name": "Classic Small",
                "dbsql_warehouse_type": "CLASSIC", "dbsql_warehouse_size": "Small",
                "dbsql_num_clusters": 1, "hours_per_month": 100,
            },
            {
                "workload_type": "DBSQL", "workload_name": "Pro Medium",
                "dbsql_warehouse_type": "PRO", "dbsql_warehouse_size": "Medium",
                "dbsql_num_clusters": 2, "hours_per_month": 200,
            },
            {
                "workload_type": "DBSQL", "workload_name": "Serverless Large",
                "dbsql_warehouse_type": "SERVERLESS", "dbsql_warehouse_size": "Large",
                "dbsql_num_clusters": 1, "hours_per_month": 730,
            },
        ])
        yield
        self.wb.close()

    def test_all_three_skus_present(self):
        sheet = self.wb.active
        skus = set()
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and 'SQL' in cell.value and 'COMPUTE' in cell.value:
                    skus.add(cell.value)
        assert 'SQL_COMPUTE' in skus, f"Missing SQL_COMPUTE in {skus}"
        assert 'SQL_PRO_COMPUTE' in skus, f"Missing SQL_PRO_COMPUTE in {skus}"
        assert 'SERVERLESS_SQL_COMPUTE' in skus, f"Missing SERVERLESS_SQL_COMPUTE in {skus}"

    def test_three_data_rows(self):
        sheet = self.wb.active
        names = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value in (
                    "Classic Small", "Pro Medium", "Serverless Large"
                ):
                    names.append(cell.value)
        assert len(names) == 3, f"Expected 3 DBSQL rows, found: {names}"

    def test_totals_row_has_sum_formulas(self):
        sheet = self.wb.active
        sum_found = False
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and 'SUM(' in cell.value.upper():
                    sum_found = True
                    break
            if sum_found:
                break
        assert sum_found, "Expected SUM formulas in totals row"

    def test_no_error_cells(self):
        sheet = self.wb.active
        error_prefixes = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A")
        errors = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(error_prefixes):
                    errors.append(f"Cell {cell.coordinate}: {cell.value}")
        assert len(errors) == 0, f"Found error cells: {errors}"
