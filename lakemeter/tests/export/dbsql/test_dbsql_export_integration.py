"""
Sprint 4: Integration test — single DBSQL Serverless export via real endpoint.

Uses FastAPI TestClient with mocked DB/auth to exercise the full export pipeline.
Multi-type integration tests are in test_dbsql_integration_multi.py.
"""
import os
import sys
import uuid
import pytest
from io import BytesIO
from unittest.mock import MagicMock

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.database import get_db
from app.auth import get_current_user
from tests.export.dbsql.conftest import make_line_item


MOCK_USER_ID = uuid.uuid4()
MOCK_ESTIMATE_ID = uuid.uuid4()


def _make_mock_user():
    user = MagicMock()
    user.user_id = MOCK_USER_ID
    user.email = "test@databricks.com"
    user.full_name = "Test User"
    user.is_active = True
    return user


def _make_mock_estimate(line_items_data):
    estimate = MagicMock()
    estimate.estimate_id = MOCK_ESTIMATE_ID
    estimate.estimate_name = "Sprint 4 DBSQL Integration Test"
    estimate.owner_user_id = MOCK_USER_ID
    estimate.is_deleted = False
    estimate.cloud = "aws"
    estimate.region = "us-east-1"
    estimate.tier = "PREMIUM"
    estimate.created_at = None
    estimate.updated_at = None
    estimate.line_items = [
        make_line_item(display_order=i, **data) for i, data in enumerate(line_items_data)
    ]
    return estimate


def _build_client_with_items(line_items_data):
    estimate = _make_mock_estimate(line_items_data)
    mock_db = MagicMock()
    mock_db.query.return_value.filter.return_value.first.return_value = estimate

    def override_db():
        yield mock_db

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: _make_mock_user()
    return TestClient(app)


def _download_excel(client) -> openpyxl.Workbook:
    url = f"/api/v1/export/estimate/{MOCK_ESTIMATE_ID}/excel"
    response = client.get(url)
    assert response.status_code == 200, f"Export failed: {response.status_code}"
    assert "spreadsheetml" in response.headers.get("content-type", "")
    return openpyxl.load_workbook(BytesIO(response.content), data_only=False)


class TestDBSQLServerlessIntegration:
    """Integration: export a DBSQL Serverless item via the real endpoint."""

    @pytest.fixture(autouse=True)
    def setup_client(self):
        self.client = _build_client_with_items([{
            "workload_type": "DBSQL",
            "workload_name": "DBSQL SL Small",
            "dbsql_warehouse_type": "SERVERLESS",
            "dbsql_warehouse_size": "Small",
            "dbsql_num_clusters": 1,
            "hours_per_month": 100,
        }])
        yield
        app.dependency_overrides.clear()

    def test_endpoint_returns_200(self):
        url = f"/api/v1/export/estimate/{MOCK_ESTIMATE_ID}/excel"
        assert self.client.get(url).status_code == 200

    def test_excel_has_serverless_sql_sku(self):
        wb = _download_excel(self.client)
        found = any(
            cell.value == "SERVERLESS_SQL_COMPUTE"
            for row in wb.active.iter_rows(values_only=False)
            for cell in row
        )
        assert found, "Expected SERVERLESS_SQL_COMPUTE SKU"
        wb.close()

    def test_serverless_row_has_zero_vm(self):
        wb = _download_excel(self.client)
        for row in wb.active.iter_rows(min_row=2, values_only=False):
            if len(row) > 3 and row[3].value == "Serverless":
                for c in range(22, 27):
                    val = row[c].value
                    assert val == 0 or val is None, \
                        f"Col {c} should be 0 for serverless, got {val}"
        wb.close()
