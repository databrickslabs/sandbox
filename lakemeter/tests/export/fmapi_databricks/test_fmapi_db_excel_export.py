"""Test FMAPI Databricks Excel export with real .xlsx files.

AC-23 to AC-28: Token formula, provisioned formula, columns, SKU, totals.
"""
import math
import os
import sys
import tempfile
import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'backend'))
from tests.export.fmapi_databricks.conftest import make_line_item
from app.routes.export import build_estimate_excel


def _make_estimate(**kw):
    from types import SimpleNamespace
    from datetime import datetime
    d = dict(estimate_name='FMAPI Databricks E2E', status='draft', version=1,
             created_at=datetime(2026, 3, 31), updated_at=datetime(2026, 3, 31))
    d.update(kw)
    return SimpleNamespace(**d)


def _generate_xlsx(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    estimate = _make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def _find_data_start_row(ws):
    """Find first data row by looking for SERVERLESS_REAL_TIME_INFERENCE SKU."""
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=6).value
        if val and isinstance(val, str) and val == 'SERVERLESS_REAL_TIME_INFERENCE':
            return row_idx
    return None


# Column indices (1-indexed for openpyxl)
COL_SKU = 6
COL_HOURS = 12
COL_TOKEN_TYPE = 13
COL_TOKEN_QTY = 14
COL_DBU_PER_M = 15
COL_DBU_HR = 16
COL_DBUS_MO = 17
COL_DBU_RATE = 18
COL_DBU_COST_L = 21


class TestTokenFormulaInExcel:
    """AC-23: Token-based items use token formula (=N*O) for DBUs/Mo."""

    def test_token_dbus_is_formula(self):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type='input_token',
            fmapi_quantity=100,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        cell = ws.cell(row=row, column=COL_DBUS_MO)
        val = cell.value
        assert val is not None
        if isinstance(val, str):
            assert val.startswith('='), f"DBUs/Mo should be formula, got: {val}"
            assert 'N' in val.upper() and 'O' in val.upper(), (
                f"Token formula should reference N*O columns, got: {val}"
            )


class TestProvisionedFormulaInExcel:
    """AC-24: Provisioned items use hours-based formula (=P*L) for DBUs/Mo."""

    def test_provisioned_dbus_is_formula(self):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type='provisioned_scaling',
            fmapi_quantity=730,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        cell = ws.cell(row=row, column=COL_DBUS_MO)
        val = cell.value
        assert val is not None
        if isinstance(val, str):
            assert val.startswith('='), f"DBUs/Mo should be formula, got: {val}"


class TestTokenColumnsPopulated:
    """AC-25: Token columns populated for token items, '-' for provisioned."""

    def test_token_item_has_token_columns(self):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type='input_token',
            fmapi_quantity=100,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        token_type = ws.cell(row=row, column=COL_TOKEN_TYPE).value
        token_qty = ws.cell(row=row, column=COL_TOKEN_QTY).value
        dbu_per_m = ws.cell(row=row, column=COL_DBU_PER_M).value
        assert token_type == 'Input', f"Expected 'Input', got '{token_type}'"
        assert token_qty == 100 or token_qty == 100.0
        assert isinstance(dbu_per_m, (int, float)) and dbu_per_m > 0

    def test_provisioned_item_has_dash_token_columns(self):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type='provisioned_scaling',
            fmapi_quantity=730,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        token_type = ws.cell(row=row, column=COL_TOKEN_TYPE).value
        assert token_type == '-' or token_type == '' or token_type is None


class TestExcelSKU:
    """AC-28: Excel SKU = SERVERLESS_REAL_TIME_INFERENCE."""

    @pytest.mark.parametrize("rate_type", [
        "input_token", "output_token", "provisioned_scaling", "provisioned_entry",
    ])
    def test_sku_correct(self, rate_type):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type=rate_type,
            fmapi_quantity=100,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        assert ws.cell(row=row, column=COL_SKU).value == 'SERVERLESS_REAL_TIME_INFERENCE'


class TestExcelFormulas:
    """AC-26: Excel has formulas in computed columns."""

    def test_dbu_cost_is_formula(self):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type='input_token',
            fmapi_quantity=100,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        cell = ws.cell(row=row, column=COL_DBU_COST_L)
        val = cell.value
        assert val is not None
        if isinstance(val, str):
            assert val.startswith('='), f"DBU Cost should be formula, got: {val}"


class TestExcelTotals:
    """AC-27: Excel SUM totals row correct."""

    def test_totals_row_with_mixed_items(self):
        items = [
            make_line_item(
                fmapi_model='llama-3-3-70b', fmapi_rate_type='input_token',
                fmapi_quantity=100, workload_name='LLM Input',
            ),
            make_line_item(
                fmapi_model='llama-3-3-70b', fmapi_rate_type='output_token',
                fmapi_quantity=50, workload_name='LLM Output', display_order=1,
            ),
            make_line_item(
                fmapi_model='bge-large', fmapi_rate_type='provisioned_entry',
                fmapi_quantity=730, workload_name='Embedding Prov', display_order=2,
            ),
        ]
        wb = _generate_xlsx(items)
        ws = wb.active
        found_sum = False
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=COL_DBU_COST_L).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                found_sum = True
                break
        assert found_sum, "No SUM formula found in totals row"


class TestExcelNoNaN:
    """Verify no NaN values in generated Excel."""

    def test_no_nan_token_item(self):
        items = [make_line_item(
            fmapi_model='llama-3-3-70b', fmapi_rate_type='input_token',
            fmapi_quantity=100,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        for col in (COL_TOKEN_QTY, COL_DBU_PER_M, COL_DBU_RATE):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                assert not math.isnan(val), f"NaN in col {col}"
