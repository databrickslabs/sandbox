"""
Sprint 3: DLT E2E Excel — Totals Row Verification Tests

Split from test_dlt_excel_e2e.py (BUG-S3-E3-1).
Generates real .xlsx with multiple DLT items, verifies SUM formulas in totals row.
"""
import os
import sys
import tempfile
import pytest
import openpyxl

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

from tests.export.dlt.conftest import make_line_item
from app.routes.export import build_estimate_excel


def _make_estimate(**kwargs):
    """Create a mock estimate object."""
    from types import SimpleNamespace
    from datetime import datetime
    defaults = {
        'estimate_name': 'DLT E2E Test Estimate',
        'status': 'draft',
        'version': 1,
        'created_at': datetime(2026, 3, 31),
        'updated_at': datetime(2026, 3, 31),
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _generate_xlsx(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    """Generate an Excel file and return the openpyxl workbook."""
    estimate = _make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def _find_totals_row(ws):
    """Find the TOTALS row."""
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value and isinstance(cell.value, str) and 'TOTALS' in cell.value:
            return row_idx
    return None


# Column mapping (1-indexed for openpyxl)
COL_DBUS_MO = 17
COL_DBU_COST_L = 21
COL_DBU_COST_D = 22
COL_DRV_VM_TOT = 25
COL_WRK_VM_TOT = 26
COL_VM_TOTAL = 27
COL_TOTAL_L = 28
COL_TOTAL_D = 29


class TestDLTExcelE2ETotals:
    """Verify the totals row uses SUM formulas."""

    @pytest.fixture
    def multi_item_workbook(self):
        """3 DLT line items for totals testing."""
        items = [
            make_line_item(
                workload_name="DLT Core Classic",
                dlt_edition="CORE",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                photon_enabled=False,
                serverless_enabled=False,
                hours_per_month=730,
            ),
            make_line_item(
                workload_name="DLT Pro Photon",
                dlt_edition="PRO",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                photon_enabled=True,
                serverless_enabled=False,
                hours_per_month=200,
            ),
            make_line_item(
                workload_name="DLT Serverless Perf",
                dlt_edition="ADVANCED",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                serverless_enabled=True,
                serverless_mode="performance",
                hours_per_month=730,
            ),
        ]
        return _generate_xlsx(items)

    def test_totals_row_exists(self, multi_item_workbook):
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        assert totals_row is not None, "TOTALS row not found"

    def test_totals_dbus_mo_has_sum(self, multi_item_workbook):
        """Col 16 (DBUs/Mo) in totals row should use SUM formula."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        cell = ws.cell(row=totals_row, column=COL_DBUS_MO)
        val = cell.value
        assert isinstance(val, str) and 'SUM' in val.upper(), \
            f"Totals DBUs/Mo should use SUM, got '{val}'"

    def test_totals_dbu_cost_list_has_sum(self, multi_item_workbook):
        """Col 20 (DBU Cost List) in totals row should use SUM."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        cell = ws.cell(row=totals_row, column=COL_DBU_COST_L)
        val = cell.value
        assert isinstance(val, str) and 'SUM' in val.upper(), \
            f"Totals DBU Cost (List) should use SUM, got '{val}'"

    def test_totals_total_cost_has_sum(self, multi_item_workbook):
        """Col 27-28 (Total Cost) in totals row should use SUM."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        for col in [COL_TOTAL_L, COL_TOTAL_D]:
            cell = ws.cell(row=totals_row, column=col)
            val = cell.value
            assert isinstance(val, str) and 'SUM' in val.upper(), \
                f"Totals Col {col} should use SUM, got '{val}'"

    def test_totals_vm_cost_has_sum(self, multi_item_workbook):
        """Cols 24-26 (VM Costs) in totals row should use SUM."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        for col in [COL_DRV_VM_TOT, COL_WRK_VM_TOT, COL_VM_TOTAL]:
            cell = ws.cell(row=totals_row, column=col)
            val = cell.value
            assert isinstance(val, str) and 'SUM' in val.upper(), \
                f"Totals Col {col} should use SUM, got '{val}'"
