"""
Sprint 2: All-Purpose Excel Export Formula Validation Tests

Builds minimal .xlsx files using the same export helper functions,
reads back with openpyxl to verify formula cells, computed values, and totals.
"""
import os
import sys
import pytest
from io import BytesIO

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

from app.routes.export import (
    _get_sku_type,
    _calculate_dbu_per_hour,
    _calculate_hours_per_month,
    _is_serverless_workload,
    _get_dbu_price,
)
from tests.export.all_purpose.conftest import make_line_item


def build_test_excel(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    """Build an Excel file from mock All-Purpose line items using export logic."""
    import xlsxwriter
    import openpyxl
    from xlsxwriter.utility import xl_col_to_name as _col

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    sheet = wb.add_worksheet('Test')

    default_fmt = wb.add_format()
    num_fmt = wb.add_format({'num_format': '#,##0'})
    dec_fmt = wb.add_format({'num_format': '#,##0.00'})
    cur_fmt = wb.add_format({'num_format': '$#,##0.00'})
    pct_fmt = wb.add_format({'num_format': '0%'})

    headers = [
        '#', 'Name', 'Type', 'Mode', 'Config', 'SKU',
        'Driver', 'Worker', 'Workers', 'DTier', 'WTier',
        'Hours/Mo', 'Token Type', 'Tokens/Mo(M)', 'DBU/1M',
        'DBU/Hr', 'DBUs/Mo', 'DBU Rate(List)', 'Discount%',
        'DBU Rate(Disc)', 'DBU Cost(List)', 'DBU Cost(Disc)',
        'Driver VM$/Hr', 'Worker VM$/Hr', 'Driver VM Cost',
        'Worker VM Cost', 'Total VM Cost', 'Total Cost(List)',
        'Total Cost(Disc)', 'Notes'
    ]
    for c, h in enumerate(headers):
        sheet.write(0, c, h)

    data_start_row = 1

    for idx, item in enumerate(line_items):
        row = data_start_row + idx
        r = row + 1

        wt = item.workload_type or 'ALL_PURPOSE'
        sku = _get_sku_type(item, cloud)
        dbu_rate, _ = _get_dbu_price(cloud, region, tier, sku)
        dbu_per_hour, warnings = _calculate_dbu_per_hour(item, cloud)
        hours = _calculate_hours_per_month(item)
        is_serverless = _is_serverless_workload(item)
        num_workers = int(item.num_workers or 1)
        total_dbus = dbu_per_hour * hours
        discount_pct = 0.0

        driver_vm_hr = 0.20 if (not is_serverless and wt in ('JOBS', 'ALL_PURPOSE', 'DLT')) else 0
        worker_vm_hr = 0.10 if (not is_serverless and wt in ('JOBS', 'ALL_PURPOSE', 'DLT')) else 0
        notes = ' | '.join(warnings) if warnings else (item.notes or '')

        sheet.write(row, 0, idx + 1)
        sheet.write(row, 1, item.workload_name or f'AP {idx+1}')
        sheet.write(row, 2, wt)
        sheet.write(row, 3, 'Serverless' if is_serverless else 'Classic')
        sheet.write(row, 4, '')
        sheet.write(row, 5, sku)
        sheet.write(row, 6, item.driver_node_type or '-')
        sheet.write(row, 7, item.worker_node_type or '-')
        sheet.write(row, 8, num_workers, num_fmt)
        sheet.write(row, 9, '-')
        sheet.write(row, 10, '-')
        sheet.write(row, 11, hours, dec_fmt)
        sheet.write(row, 12, '-')
        sheet.write(row, 13, '-')
        sheet.write(row, 14, '-')
        sheet.write(row, 15, dbu_per_hour, dec_fmt)

        # Col 16: DBUs/Mo = DBU/Hr x Hours/Mo
        formula = f'={_col(15)}{r}*{_col(11)}{r}'
        sheet.write_formula(row, 16, formula, num_fmt, total_dbus)

        # Col 17: DBU Rate (List)
        sheet.write(row, 17, dbu_rate, cur_fmt)
        # Col 18: Discount %
        sheet.write(row, 18, discount_pct, pct_fmt)

        # Col 19: DBU Rate (Disc.)
        disc_rate = dbu_rate * (1 - discount_pct)
        formula = f'={_col(17)}{r}*(1-{_col(18)}{r})'
        sheet.write_formula(row, 19, formula, cur_fmt, disc_rate)

        # Col 20: DBU Cost (List) = Q*R
        dbu_cost_list = total_dbus * dbu_rate
        formula = f'={_col(16)}{r}*{_col(17)}{r}'
        sheet.write_formula(row, 20, formula, cur_fmt, dbu_cost_list)

        # Col 21: DBU Cost (Disc.)
        dbu_cost_disc = total_dbus * disc_rate
        formula = f'={_col(16)}{r}*{_col(19)}{r}'
        sheet.write_formula(row, 21, formula, cur_fmt, dbu_cost_disc)

        if is_serverless:
            for c in range(22, 27):
                sheet.write(row, c, 0, cur_fmt)
        else:
            sheet.write(row, 22, driver_vm_hr, cur_fmt)
            sheet.write(row, 23, worker_vm_hr, cur_fmt)
            driver_vm_total = driver_vm_hr * hours
            formula = f'={_col(22)}{r}*{_col(11)}{r}'
            sheet.write_formula(row, 24, formula, cur_fmt, driver_vm_total)
            worker_vm_total = worker_vm_hr * hours * num_workers
            formula = f'={_col(23)}{r}*{_col(11)}{r}*{_col(8)}{r}'
            sheet.write_formula(row, 25, formula, cur_fmt, worker_vm_total)
            formula = f'={_col(24)}{r}+{_col(25)}{r}'
            sheet.write_formula(row, 26, formula, cur_fmt, driver_vm_total + worker_vm_total)

        vm_total = 0
        if not is_serverless:
            vm_total = driver_vm_hr * hours + worker_vm_hr * hours * num_workers
        formula = f'={_col(20)}{r}+{_col(26)}{r}'
        sheet.write_formula(row, 27, formula, cur_fmt, dbu_cost_list + vm_total)
        formula = f'={_col(21)}{r}+{_col(26)}{r}'
        sheet.write_formula(row, 28, formula, cur_fmt, dbu_cost_disc + vm_total)

        sheet.write(row, 29, notes)

    data_end_row = data_start_row + len(line_items) - 1
    totals_row = data_end_row + 2
    ds = data_start_row + 1
    de = data_end_row + 1
    sheet.write(totals_row, 0, 'TOTALS')
    sheet.write_formula(totals_row, 16, f'=SUM({_col(16)}{ds}:{_col(16)}{de})', num_fmt)
    sheet.write_formula(totals_row, 20, f'=SUM({_col(20)}{ds}:{_col(20)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 21, f'=SUM({_col(21)}{ds}:{_col(21)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 24, f'=SUM({_col(24)}{ds}:{_col(24)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 25, f'=SUM({_col(25)}{ds}:{_col(25)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 26, f'=SUM({_col(26)}{ds}:{_col(26)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 27, f'=SUM({_col(27)}{ds}:{_col(27)}{de})', cur_fmt)
    sheet.write_formula(totals_row, 28, f'=SUM({_col(28)}{ds}:{_col(28)}{de})', cur_fmt)

    wb.close()
    output.seek(0)

    oxl_wb = openpyxl.load_workbook(output, data_only=False)
    oxl_sheet = oxl_wb.active
    return oxl_wb, oxl_sheet, data_start_row, totals_row


# ============================================================
# Test: All-Purpose Classic Formula Presence
# ============================================================

class TestAllPurposeClassicFormulas:
    """Verify formula cells for All-Purpose Classic Standard export."""

    @pytest.fixture
    def classic_excel(self):
        item = make_line_item(
            workload_type="ALL_PURPOSE",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=4, photon_enabled=False, serverless_enabled=False,
            hours_per_month=730,
        )
        wb, sheet, ds, totals = build_test_excel([item])
        yield sheet, ds, totals
        wb.close()

    def test_dbus_mo_formula(self, classic_excel):
        """Col 16 (DBUs/Mo) should be a formula."""
        sheet, ds, _ = classic_excel
        cell = sheet.cell(row=ds + 1, column=17)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_dbu_cost_list_formula(self, classic_excel):
        """Col 20 (DBU Cost List) should be a formula."""
        sheet, ds, _ = classic_excel
        cell = sheet.cell(row=ds + 1, column=21)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_driver_vm_cost_formula(self, classic_excel):
        """Col 24 (Driver VM Cost) should be a formula for classic."""
        sheet, ds, _ = classic_excel
        cell = sheet.cell(row=ds + 1, column=25)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_worker_vm_cost_formula(self, classic_excel):
        """Col 25 (Worker VM Cost) should be a formula for classic."""
        sheet, ds, _ = classic_excel
        cell = sheet.cell(row=ds + 1, column=26)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_total_vm_cost_formula(self, classic_excel):
        """Col 26 (Total VM Cost) should be a formula."""
        sheet, ds, _ = classic_excel
        cell = sheet.cell(row=ds + 1, column=27)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_total_cost_formula(self, classic_excel):
        """Col 27 (Total Cost List) should be a formula."""
        sheet, ds, _ = classic_excel
        cell = sheet.cell(row=ds + 1, column=28)
        assert isinstance(cell.value, str) and cell.value.startswith('=')


# ============================================================
# Test: All-Purpose Serverless — No VM formulas
# ============================================================

class TestAllPurposeServerlessFormulas:
    """Verify serverless export has no VM cost formulas."""

    @pytest.fixture
    def serverless_excel(self):
        item = make_line_item(
            workload_type="ALL_PURPOSE",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=4, serverless_enabled=True,
            serverless_mode="performance", hours_per_month=730,
        )
        wb, sheet, ds, totals = build_test_excel([item])
        yield sheet, ds, totals
        wb.close()

    def test_dbus_mo_formula(self, serverless_excel):
        """Col 16 (DBUs/Mo) should still be a formula."""
        sheet, ds, _ = serverless_excel
        cell = sheet.cell(row=ds + 1, column=17)
        assert isinstance(cell.value, str) and cell.value.startswith('=')

    def test_vm_cost_columns_zero(self, serverless_excel):
        """VM cost columns (22-26) should be 0 for serverless."""
        sheet, ds, _ = serverless_excel
        for col in [23, 24, 25, 26, 27]:  # openpyxl 1-indexed = cols 22-26
            cell = sheet.cell(row=ds + 1, column=col)
            assert cell.value == 0, f"Col {col-1} should be 0 for serverless, got {cell.value}"


# ============================================================
# Test: Totals Row SUM Formulas
# ============================================================

class TestAllPurposeTotalsRow:
    """Verify SUM formulas in the totals row."""

    @pytest.fixture
    def multi_item_excel(self):
        items = [
            make_line_item(
                workload_type="ALL_PURPOSE",
                workload_name="AP Classic",
                driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                num_workers=4, photon_enabled=False, serverless_enabled=False,
                hours_per_month=730,
            ),
            make_line_item(
                workload_type="ALL_PURPOSE",
                workload_name="AP Photon",
                driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                num_workers=4, photon_enabled=True, serverless_enabled=False,
                hours_per_month=730,
            ),
            make_line_item(
                workload_type="ALL_PURPOSE",
                workload_name="AP Serverless",
                driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
                num_workers=4, serverless_enabled=True,
                serverless_mode="performance", hours_per_month=730,
            ),
        ]
        wb, sheet, ds, totals = build_test_excel(items)
        yield sheet, ds, totals
        wb.close()

    def test_totals_dbus_mo_sum(self, multi_item_excel):
        """Totals DBUs/Mo should be a SUM formula."""
        sheet, _, totals = multi_item_excel
        cell = sheet.cell(row=totals + 1, column=17)
        assert isinstance(cell.value, str) and cell.value.startswith('=SUM')

    def test_totals_dbu_cost_sum(self, multi_item_excel):
        """Totals DBU Cost (List) should be a SUM formula."""
        sheet, _, totals = multi_item_excel
        cell = sheet.cell(row=totals + 1, column=21)
        assert isinstance(cell.value, str) and cell.value.startswith('=SUM')

    def test_totals_total_cost_sum(self, multi_item_excel):
        """Totals Total Cost (List) should be a SUM formula."""
        sheet, _, totals = multi_item_excel
        cell = sheet.cell(row=totals + 1, column=28)
        assert isinstance(cell.value, str) and cell.value.startswith('=SUM')

    def test_totals_vm_cost_sum(self, multi_item_excel):
        """Totals VM Cost should be a SUM formula."""
        sheet, _, totals = multi_item_excel
        cell = sheet.cell(row=totals + 1, column=27)
        assert isinstance(cell.value, str) and cell.value.startswith('=SUM')


# ============================================================
# Test: Computed Values Match Expected
# ============================================================

class TestAllPurposeComputedValues:
    """Verify computed cached values match expected calculations."""

    def test_classic_4workers_730hrs(self):
        """DBU/hr=5, hours=730, total_dbus=3650."""
        item = make_line_item(
            workload_type="ALL_PURPOSE",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=4, photon_enabled=False, serverless_enabled=False,
            hours_per_month=730,
        )
        dbu_hr, _ = _calculate_dbu_per_hour(item, "aws")
        hours = _calculate_hours_per_month(item)
        assert dbu_hr == pytest.approx(5.0)
        assert hours == pytest.approx(730.0)
        assert dbu_hr * hours == pytest.approx(3650.0)

    def test_photon_4workers_730hrs(self):
        """Photon: DBU/hr=10, hours=730, total_dbus=7300."""
        item = make_line_item(
            workload_type="ALL_PURPOSE",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=4, photon_enabled=True, serverless_enabled=False,
            hours_per_month=730,
        )
        dbu_hr, _ = _calculate_dbu_per_hour(item, "aws")
        assert dbu_hr == pytest.approx(10.0)
        assert dbu_hr * 730 == pytest.approx(7300.0)

    def test_serverless_perf_4workers_730hrs(self):
        """Serverless perf: DBU/hr=20, hours=730, total_dbus=14600."""
        item = make_line_item(
            workload_type="ALL_PURPOSE",
            driver_node_type="i3.xlarge", worker_node_type="i3.xlarge",
            num_workers=4, serverless_enabled=True,
            serverless_mode="performance", hours_per_month=730,
        )
        dbu_hr, _ = _calculate_dbu_per_hour(item, "aws")
        assert dbu_hr == pytest.approx(20.0)
        assert dbu_hr * 730 == pytest.approx(14600.0)

    def test_run_based_hours_calculation(self):
        """5 runs x 45 min x 20 days = 75 hours."""
        item = make_line_item(
            workload_type="ALL_PURPOSE",
            runs_per_day=5, avg_runtime_minutes=45, days_per_month=20,
        )
        hours = _calculate_hours_per_month(item)
        assert hours == pytest.approx(75.0)
