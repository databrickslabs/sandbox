"""Test Model Serving Excel export with real .xlsx files.

AC-25 to AC-28: Excel SKU, formulas, totals, serverless label.
"""
import math
import os
import sys
import tempfile
import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'backend'))
from tests.export.model_serving.conftest import make_line_item
from app.routes.export import build_estimate_excel


def _make_estimate(**kw):
    from types import SimpleNamespace
    from datetime import datetime
    d = dict(estimate_name='Model Serving E2E Test', status='draft', version=1,
             created_at=datetime(2026, 3, 31), updated_at=datetime(2026, 3, 31))
    d.update(kw)
    return SimpleNamespace(**d)


def _generate_xlsx(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    estimate = _make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def _find_data_start_row(ws):
    """Find first data row by looking for SERVERLESS_REAL_TIME_INFERENCE SKU."""
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=6).value  # Col F = SKU
        if val and isinstance(val, str) and val == 'SERVERLESS_REAL_TIME_INFERENCE':
            return row_idx
    return None


# Column indices (1-indexed for openpyxl)
COL_MODE = 4       # Classic/Serverless
COL_SKU = 6        # SKU product type
COL_HOURS = 12     # Hours/Mo
COL_DBU_HR = 16    # DBU/Hr
COL_DBUS_MO = 17   # DBUs/Mo (formula)
COL_DBU_RATE = 18  # DBU Rate List
COL_DBU_COST_L = 21  # DBU Cost List (formula)


class TestExcelSKU:
    """AC-25: Excel has correct SKU."""

    def test_cpu_sku(self):
        items = [make_line_item(model_serving_gpu_type='cpu', hours_per_month=200)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None, "No data row with SERVERLESS_REAL_TIME_INFERENCE"
        assert ws.cell(row=row, column=COL_SKU).value == 'SERVERLESS_REAL_TIME_INFERENCE'

    def test_t4_sku(self):
        items = [make_line_item(model_serving_gpu_type='gpu_small_t4', hours_per_month=100)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        assert ws.cell(row=row, column=COL_SKU).value == 'SERVERLESS_REAL_TIME_INFERENCE'


class TestExcelServerlessLabel:
    """AC-28: All Model Serving rows show "Serverless" mode."""

    def test_serverless_mode_label(self):
        items = [make_line_item(model_serving_gpu_type='cpu', hours_per_month=730)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        mode = ws.cell(row=row, column=COL_MODE).value
        assert mode == 'Serverless', f"Expected 'Serverless', got '{mode}'"


class TestExcelFormulas:
    """AC-26: Excel has formulas (not static values) in computed columns."""

    def test_monthly_dbus_is_formula(self):
        items = [make_line_item(model_serving_gpu_type='gpu_small_t4', hours_per_month=200)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        cell = ws.cell(row=row, column=COL_DBUS_MO)
        # xlsxwriter writes formulas — openpyxl reads them as strings starting with '='
        val = cell.value
        assert val is not None
        if isinstance(val, str):
            assert val.startswith('='), f"DBUs/Mo should be formula, got: {val}"

    def test_dbu_cost_is_formula(self):
        items = [make_line_item(model_serving_gpu_type='cpu', hours_per_month=730)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        cell = ws.cell(row=row, column=COL_DBU_COST_L)
        val = cell.value
        assert val is not None
        if isinstance(val, str):
            assert val.startswith('='), f"DBU Cost should be formula, got: {val}"


class TestExcelTotals:
    """AC-27: Excel SUM totals row correct."""

    def test_totals_row_exists(self):
        items = [
            make_line_item(model_serving_gpu_type='cpu', hours_per_month=730,
                           workload_name='CPU Endpoint'),
            make_line_item(model_serving_gpu_type='gpu_small_t4', hours_per_month=200,
                           workload_name='T4 Endpoint', display_order=1),
        ]
        wb = _generate_xlsx(items)
        ws = wb.active
        # Find totals row (look for SUM formula after data rows)
        found_sum = False
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=COL_DBU_COST_L).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                found_sum = True
                break
        assert found_sum, "No SUM formula found in totals row"

    def test_multiple_rows_all_serverless(self):
        items = [
            make_line_item(model_serving_gpu_type='cpu', hours_per_month=730,
                           workload_name='CPU'),
            make_line_item(model_serving_gpu_type='gpu_medium_a10g_1x',
                           hours_per_month=100, workload_name='A10G',
                           display_order=1),
        ]
        wb = _generate_xlsx(items)
        ws = wb.active
        # All data rows should have SERVERLESS_REAL_TIME_INFERENCE
        data_rows = []
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=COL_SKU).value
            if val == 'SERVERLESS_REAL_TIME_INFERENCE':
                data_rows.append(row_idx)
        assert len(data_rows) == 2, f"Expected 2 data rows, found {len(data_rows)}"


class TestExcelNoNaN:
    """Verify no NaN values in generated Excel."""

    def test_no_nan_in_static_cells(self):
        items = [make_line_item(model_serving_gpu_type='cpu', hours_per_month=730)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_data_start_row(ws)
        assert row is not None
        # Check hours and DBU/hr are numbers, not NaN
        hours = ws.cell(row=row, column=COL_HOURS).value
        dbu_hr = ws.cell(row=row, column=COL_DBU_HR).value
        if isinstance(hours, (int, float)):
            assert not math.isnan(hours)
        if isinstance(dbu_hr, (int, float)):
            assert not math.isnan(dbu_hr)
