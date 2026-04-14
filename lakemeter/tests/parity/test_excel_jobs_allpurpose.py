"""End-to-end Excel generation parity tests for JOBS and ALL_PURPOSE workloads.

Generates actual Excel workbooks and verifies that cell values match
frontend calculation expectations.
"""
import os
import tempfile
import pytest
import openpyxl
from types import SimpleNamespace
from datetime import datetime

from .conftest import make_item
from .frontend_calc import (
    fe_compute_dbu_per_hour, fe_hours_per_month, fe_monthly_dbu_cost,
)

CLOUD = 'aws'
REGION = 'us-east-1'
TIER = 'PREMIUM'
TOL = 0.01

# Column indices (1-indexed for openpyxl, matching excel_row_writer.py)
COL_SKU = 6
COL_HOURS = 12
COL_DBU_HR = 16
COL_DBUS_MO = 17
COL_DBU_RATE = 18
COL_DBU_COST_L = 21
COL_TOTAL_L = 28


def _make_estimate():
    return SimpleNamespace(
        estimate_name='Parity Test', status='draft', version=1,
        created_at=datetime(2026, 4, 1), updated_at=datetime(2026, 4, 1),
    )


def _generate_xlsx(items):
    """Generate Excel and return openpyxl Workbook."""
    from app.routes.export import build_estimate_excel
    output = build_estimate_excel(_make_estimate(), items, CLOUD, REGION, TIER)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    os.unlink(tmp_path)
    return wb


def _find_data_row(ws, sku_value):
    """Find first data row with the given SKU."""
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=COL_SKU).value == sku_value:
            return row_idx
    return None


class TestJobsClassicExcel:
    """End-to-end: JOBS Classic item → Excel cell values."""

    def test_cells_match_frontend(self, pricing):
        driver_dbu = pricing['instance_dbu_rates']['aws']['m5d.xlarge']['dbu_rate']
        worker_dbu = pricing['instance_dbu_rates']['aws']['i3.xlarge']['dbu_rate']
        item = make_item(
            workload_type='JOBS', workload_name='Jobs Classic Test',
            driver_node_type='m5d.xlarge', worker_node_type='i3.xlarge',
            num_workers=2, hours_per_month=100,
        )
        wb = _generate_xlsx([item])
        ws = wb.active
        row = _find_data_row(ws, 'JOBS_COMPUTE')
        assert row is not None, "JOBS_COMPUTE row not found in Excel"

        fe_dbu_hr = fe_compute_dbu_per_hour(
            driver_dbu_rate=driver_dbu, worker_dbu_rate=worker_dbu,
            num_workers=2, workload_type='JOBS',
        )
        fe_hours = 100.0
        fe_dbus_mo = fe_dbu_hr * fe_hours

        assert ws.cell(row=row, column=COL_DBU_HR).value == pytest.approx(fe_dbu_hr, abs=TOL)
        assert ws.cell(row=row, column=COL_HOURS).value == pytest.approx(fe_hours, abs=TOL)
        assert ws.cell(row=row, column=COL_DBUS_MO).value == pytest.approx(fe_dbus_mo, abs=TOL)


class TestJobsServerlessPerformanceExcel:
    """End-to-end: JOBS Serverless Performance → Excel cell values."""

    def test_cells_match_frontend(self, pricing):
        driver_dbu = pricing['instance_dbu_rates']['aws']['m5d.xlarge']['dbu_rate']
        worker_dbu = pricing['instance_dbu_rates']['aws']['i3.xlarge']['dbu_rate']
        photon_mult = pricing['dbu_multipliers']['aws:JOBS_COMPUTE:photon']['multiplier']
        item = make_item(
            workload_type='JOBS', workload_name='Jobs Serverless Perf',
            driver_node_type='m5d.xlarge', worker_node_type='i3.xlarge',
            num_workers=2, serverless_enabled=True,
            serverless_mode='performance', hours_per_month=50,
        )
        wb = _generate_xlsx([item])
        ws = wb.active
        row = _find_data_row(ws, 'JOBS_SERVERLESS_COMPUTE')
        assert row is not None, "JOBS_SERVERLESS_COMPUTE row not found"

        fe_dbu_hr = fe_compute_dbu_per_hour(
            driver_dbu_rate=driver_dbu, worker_dbu_rate=worker_dbu,
            num_workers=2, serverless_enabled=True,
            serverless_mode='performance', workload_type='JOBS',
            photon_multiplier=photon_mult,
        )
        fe_hours = 50.0
        fe_dbus_mo = fe_dbu_hr * fe_hours

        assert ws.cell(row=row, column=COL_DBU_HR).value == pytest.approx(fe_dbu_hr, abs=TOL)
        assert ws.cell(row=row, column=COL_HOURS).value == pytest.approx(fe_hours, abs=TOL)
        assert ws.cell(row=row, column=COL_DBUS_MO).value == pytest.approx(fe_dbus_mo, abs=TOL)


class TestAllPurposeServerlessExcel:
    """End-to-end: ALL_PURPOSE Serverless → Excel cell values."""

    def test_cells_match_frontend(self, pricing):
        driver_dbu = pricing['instance_dbu_rates']['aws']['m5d.xlarge']['dbu_rate']
        worker_dbu = pricing['instance_dbu_rates']['aws']['i3.xlarge']['dbu_rate']
        photon_mult = pricing['dbu_multipliers']['aws:ALL_PURPOSE_COMPUTE:photon']['multiplier']
        item = make_item(
            workload_type='ALL_PURPOSE', workload_name='AP Serverless',
            driver_node_type='m5d.xlarge', worker_node_type='i3.xlarge',
            num_workers=1, serverless_enabled=True, hours_per_month=300,
        )
        wb = _generate_xlsx([item])
        ws = wb.active
        row = _find_data_row(ws, 'ALL_PURPOSE_SERVERLESS_COMPUTE')
        assert row is not None, "ALL_PURPOSE_SERVERLESS_COMPUTE row not found"

        fe_dbu_hr = fe_compute_dbu_per_hour(
            driver_dbu_rate=driver_dbu, worker_dbu_rate=worker_dbu,
            num_workers=1, serverless_enabled=True,
            workload_type='ALL_PURPOSE', photon_multiplier=photon_mult,
        )
        fe_hours = 300.0
        fe_dbus_mo = fe_dbu_hr * fe_hours

        assert ws.cell(row=row, column=COL_DBU_HR).value == pytest.approx(fe_dbu_hr, abs=TOL)
        assert ws.cell(row=row, column=COL_HOURS).value == pytest.approx(fe_hours, abs=TOL)
        assert ws.cell(row=row, column=COL_DBUS_MO).value == pytest.approx(fe_dbus_mo, abs=TOL)


class TestExcelFormulaConsistency:
    """Verify Excel cached values match formula expectations."""

    def test_dbu_cost_equals_dbus_times_rate(self, pricing):
        """Col 20 (DBU Cost List) = Col 16 (DBUs/Mo) × Col 17 (DBU Rate)."""
        item = make_item(
            workload_type='JOBS', workload_name='Formula Check',
            driver_node_type='m5d.xlarge', worker_node_type='i3.xlarge',
            num_workers=4, hours_per_month=200,
        )
        wb = _generate_xlsx([item])
        ws = wb.active
        row = _find_data_row(ws, 'JOBS_COMPUTE')
        assert row is not None

        dbus_mo = ws.cell(row=row, column=COL_DBUS_MO).value
        dbu_rate = ws.cell(row=row, column=COL_DBU_RATE).value
        dbu_cost = ws.cell(row=row, column=COL_DBU_COST_L).value

        assert dbu_cost == pytest.approx(dbus_mo * dbu_rate, abs=TOL)

    def test_total_cost_equals_dbu_cost_for_serverless(self, pricing):
        """For serverless, Total Cost = DBU Cost (no VM costs)."""
        item = make_item(
            workload_type='JOBS', workload_name='Serverless Total',
            driver_node_type='m5d.xlarge', worker_node_type='i3.xlarge',
            num_workers=2, serverless_enabled=True,
            serverless_mode='standard', hours_per_month=100,
        )
        wb = _generate_xlsx([item])
        ws = wb.active
        row = _find_data_row(ws, 'JOBS_SERVERLESS_COMPUTE')
        assert row is not None

        dbu_cost = ws.cell(row=row, column=COL_DBU_COST_L).value
        total_cost = ws.cell(row=row, column=COL_TOTAL_L).value
        assert total_cost == pytest.approx(dbu_cost, abs=TOL)
